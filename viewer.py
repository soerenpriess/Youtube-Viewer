import psutil
import time
import keyboard
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import requests
from selenium import webdriver
import re
import random

useXlxs = False
url = input("URL: ") #Url of the Video
iteration = input("Iterations (1 Iteration = 30-40 Tabs): ") #how many iterations (1 Iteration = 30-40 repetitions)
killTime = input("Kill-Time: ") #how longe the video should play after every tab is open
startViews = input("Current-Views: ") #current views of the video
beginTime = datetime.now() #start time


def checkFreeSources():
    """
    check if free memory and cpu is > 15 then break the loop and returns True. If not wait 5sec and check again
    :return: True
    """
    while True:
        memoryFree = psutil.virtual_memory().available * 100 / psutil.virtual_memory().total
        cpuFree = 100 - psutil.cpu_percent()
        print("Free Memory: " + str(memoryFree) + "%")
        print("Free Cpu: " + str(cpuFree) + "%")

        if memoryFree > 15 and cpuFree > 15:
            break
        else:
            time.sleep(5)

    return True


def kill():
    """
    close the tabs with ctrl+w
    :return: /
    """
    keyboard.press_and_release('ctrl+w')  # closes the last tab
    print("killed")



def startBot(currIteration, proxy):
    """
    opens 30-40 times a chrome tab with the video url and close them after the killTime again
    :param currIteration: current Iteration
    :param proxy: proxy to use
    :return: True
    """
    proxy = str(proxy)
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--proxy-server=http://%s' % proxy) # set the proxy for chrome
    chrome_options.add_argument('--window-size=640,480') # set the window size to use less ressources
    chrome_options.add_argument('--no-sandbox') # no sandbox mode
    chrome_options.add_experimental_option("excludeSwitches", ['enable-automation']) # disable infobar
    chrome = webdriver.Chrome(executable_path="./chromedriver.exe", options=chrome_options)
    chrome.implicitly_wait(0.6)
    rounds = random.randrange(30, 40)
    for x in range(rounds):
        if checkFreeSources():
            try:
                chrome.get(url)
                time.sleep(0.5)
                keyboard.press_and_release('space')  # start the video
                if x != rounds:
                    chrome.execute_script("window.open('about:blank', 'tab" + str(x + 1) + "');")
                    chrome.switch_to.window("tab" + str(x + 1))
            except IOError as e:
                print(e)

            print("Step " + str(x+1) + " from " + str(rounds) + " | Iteration: " + str(currIteration+1) + " from " + str(iteration))
            time.sleep(1.5)
        else:
            print("error exit")

    time.sleep((int(killTime)))

    for x in range(40):
        kill()
        time.sleep(0.5)

    return True


def main():
    """
    call for each iteration the required functions to get a new proxy and open the tabs.
    after that the params will be saved in xlxs
    :return:
    """
    for x in range(int(iteration)):
        print("Start Iteration " + str(x))
        work = testProxies(getProxyList())  # stores working proxies
        print(work)  # shows the current used proxiy
        startBot(x, work[0])

    endTime = datetime.now()
    runTime = endTime - beginTime
    endViews = input("End-Views: ")
    if useXlxs:
        setXlsxData(str(endTime), str(runTime), str(endViews))
    print("Run-Time: " + str(runTime))


def getXlsxData():
    """
    get the last used row from the xlsx
    :return: last used row number
    """
    path = "test.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    maxRow = sheet_obj.max_row
    print(str(maxRow))
    return maxRow



def setXlsxData(endTime, runTime, endViews):
    """
    saves requiered data in xlxs data
    :param endTime: endtime
    :param runTime: total runtime
    :param endViews: endviews
    :return: /
    """
    wb = load_workbook("test.xlsx")
    sheet = wb.active
    nextRow = (int(getXlsxData()) + 1)
    columns = sheet["A" + str(nextRow)].value = nextRow-1
    columns = sheet["B" + str(nextRow)].value = beginTime
    columns = sheet["C" + str(nextRow)].value = endTime
    columns = sheet["D" + str(nextRow)].value = runTime
    columns = sheet["E" + str(nextRow)].value = iteration
    columns = sheet["F" + str(nextRow)].value = url
    columns = sheet["G" + str(nextRow)].value = startViews
    columns = sheet["H" + str(nextRow)].value = endViews
    wb.save("test.xlsx")



def getProxyList():
    """
    gets a list of free proxy from the url and format it
    :return: list with proxies
    """
    url = 'https://free-proxy-list.net/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Cafari/537.36'}
    source = str(requests.get(url, headers=headers, timeout=10).text)
    data = [list(filter(None, i))[0] for i in re.findall('<td class="hm">(.*?)</td>|<td>(.*?)</td>', source)]
    groupings = [dict(zip(['ip', 'port', 'code', 'using_anonymous'], data[i:i + 4])) for i in range(0, len(data), 4)]
    final_groupings = [{'full_ip': "{ip}:{port}".format(**i)} for i in groupings]
    return final_groupings



def testProxies(proxies):
    """
    gets a list of proxies and checks each one to see if the ip address is returned when the url is called.
    If so, the proxy is saved and returned.
    If not, the next one will be tested
    :param proxies: list of proxies
    :return: list with working proxy
    """
    workingProxy = []
    for i in range(len(proxies)):
        try:
            r = requests.get('https://httpbin.org/ip', proxies={'http': "http://" + proxies[i]['full_ip'], 'https': "https://" + proxies[i]['full_ip']}, timeout=1)
        except IOError:
            print("Proxy    " + proxies[i]['full_ip'] + "    doesn't work!")
        else:
            workingProxy.append(proxies[i]['full_ip'])
            print("Proxy    " + proxies[i]['full_ip'] + "    work!")
            return workingProxy

    print("Found no Proxy!")

main()




